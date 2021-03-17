#importing Modules And Giving A Name To It 
import tkinter as tk
import shutil
import xlwt
import xlrd
import os
import openpyxl
import pyexcel
from xlutils.copy import copy
from xlutils.styles import Styles
from tkinter import ttk
from tkinter import filedialog
from tkinter import messagebox as msg
from PIL import Image,ImageTk
from xlwt import Workbook,XFStyle,Borders,Pattern,Font,Alignment


#Functions
def table_contents():
    table.delete(*table.get_children())
    try:
        l=0
        sheet=pyexcel.get_sheet(file_name="Database.xls")
        for i in sheet:
            if i=="-" or i=="+":
                pass
            else:
                j=tuple(i)
                k=list(j)
                if l!=0:
                    table.insert("","end",values=k)
                else:
                    l+=1
    except FileNotFoundError:
        pass

def get_cursor(values):
    cursor=table.focus()
    contents=table.item(cursor)
    row_contents=contents["values"]
    fst_name.set(row_contents[1])
    initial.set(row_contents[2])
    dob=row_contents[3].split(" ")
    dob_date.set(dob[0])
    dob_month.set(dob[1])
    dob_year.set(dob[2])
    age.set(row_contents[4])
    gender.set(row_contents[5])
    father_name.set(row_contents[6])
    father_mobile.set(row_contents[7])
    mother_name.set(row_contents[8])
    mother_mobile.set(row_contents[9])
    whatsapp_number.set(row_contents[10])
    add=row_contents[11].split(" ")
    house_no.set(add[0])
    string_address=" ".join(add[1:])
    address.set(string_address)
    district.set(row_contents[12])
    pin_code.set(row_contents[13])
    state.set(row_contents[14])
    father_mail.set(row_contents[15])
    student_mail.set(row_contents[16])
    group.set(row_contents[17])
    section.set(row_contents[18])
    app_no.set(row_contents[19])  
    
def image_ftn():
    global image_display_lbl
    app_no_submit_image=app_no.get()
    image_no=app_no.get()
    if image_no=="":
        msg.showwarning("Warning","Please Fill The Application Number And Proceed To Add Image")
    else:

        image_path=filedialog.askopenfilename(title="Select A Image",filetype=(("jpeg","*.jpg"),("All Files","*.*")))
        try:
            application_no=0
            try:
                excel_file_image=xlrd.open_workbook("Database.xls")
                column=19
                app_no_submit_checking0=int(app_no_submit_image)
                app_no_submit_checking=app_no_submit_checking0/1
                for sheet in excel_file_image.sheets():
                    for row_in_sheet in range(sheet.nrows):
                        for column_in_sheet in range(sheet.ncols):
                            if column_in_sheet==column:
                                app_no_excel=sheet.cell(row_in_sheet,column_in_sheet).value
                                if app_no_submit_checking==app_no_excel:
                                    application_no+=1
                                    msg.showwarning("Warning","Application Number Already Exists")
            except FileNotFoundError:
                pass
            if application_no==0:
                std_image=shutil.copy(image_path,"Students Image\profile "+image_no+".jpg")
                msg.showinfo("Information","You Can Preview The Student Image By Clicking\nStudents Image >> profile (Application Number) ")


        except FileNotFoundError:
            os.makedirs("Students Image")
            shutil.copy(image_path,"Students Image\profile "+image_no+".jpg")
            msg.showinfo("Information","Yoy Can Preview The Student Image By Clicking\nStudents Image >> profile (Application Number) ")
            

        

                                 
def exit_software():
    exit_answer=msg.askyesnocancel("Quit","Are You Sure You Want To Quit ?\nIf You Haven't Saved Your Data It Will Be Lost")
    if exit_answer==True:
        database.quit()
        database.destroy()
    else:
        pass

def clearing():
    fst_name.set("")
    initial.set("")
    dob_date.set("")
    dob_month.set("")
    dob_year.set("")
    age.set("")
    gender.set("")
    father_name.set("")
    father_mobile.set("+91 ")
    mother_name.set("")
    mother_mobile.set("+91 ")
    whatsapp_number.set("+91 ")
    house_no.set("")
    address.set("")
    district.set("")
    pin_code.set("")
    state.set("")
    father_mail.set("")
    student_mail.set("")
    group.set("")
    section.set("")
    app_no.set("") 
    
def clear():
    clear_ans=msg.askyesno("Warning","The Data Entered Will Be Lost\nAre You Sure You Wish To Do This?")
    if clear_ans==True:
        clearing()
        return

def update():
    cursor=table.focus()
    contents=table.item(cursor)
    row_contents=contents["values"]
    
    fst_name_submit=(fst_name.get())
    initial_submit=initial.get()
    dob_date_submit=dob_date.get()
    dob_month_submit=dob_month.get()
    dob_year_submit=dob_year.get()
    age_submit=age.get()
    gender_submit=gender.get()
    father_name_submit=father_name.get()
    father_mobile_submit=father_mobile.get()
    mother_name_submit=mother_name.get()
    mother_mobile_submit=mother_mobile.get()
    whatsapp_number_submit=whatsapp_number.get()
    house_no_submit=house_no.get()
    address_submit=address.get()
    district_submit=district.get()
    pin_code_submit=pin_code.get()
    state_submit=state.get()
    father_mail_submit=father_mail.get()
    student_mail_submit=student_mail.get()
    group_submit=group.get()
    section_submit=section.get()
    app_no_submit=app_no.get()

    try:
        if str(app_no_submit)==str(row_contents[19]):
            update_value=msg.askyesno("Warning","Are You Sure You Wate To Update This Student Details")
            if update_value==True:
                row=row_contents[0]
                excel_file_read_write=xlrd.open_workbook("Database.xls")
                excel_file_write=copy(excel_file_read_write)
                excel_file_write_sheet=excel_file_write.get_sheet(0)
                excel_file_write_sheet.write(row,0,row)
                excel_file_write_sheet.write(row,1,fst_name_submit)
                excel_file_write_sheet.write(row,2,initial_submit)
                excel_file_write_sheet.write(row,3,dob_date_submit+" "+dob_month_submit+" "+dob_year_submit)
                excel_file_write_sheet.write(row,4,age_submit)
                excel_file_write_sheet.write(row,5,gender_submit)
                excel_file_write_sheet.write(row,6,father_name_submit)
                excel_file_write_sheet.write(row,7,father_mobile_submit)
                excel_file_write_sheet.write(row,8,mother_name_submit)
                excel_file_write_sheet.write(row,9,mother_mobile_submit)
                excel_file_write_sheet.write(row,10,whatsapp_number_submit)
                excel_file_write_sheet.write(row,11,house_no_submit+" "+address_submit)
                excel_file_write_sheet.write(row,12,district_submit)
                excel_file_write_sheet.write(row,13,pin_code_submit)
                excel_file_write_sheet.write(row,14,state_submit)
                excel_file_write_sheet.write(row,15,father_mail_submit)
                excel_file_write_sheet.write(row,16,student_mail_submit)
                excel_file_write_sheet.write(row,17,group_submit)
                excel_file_write_sheet.write(row,18,section_submit)
                excel_file_write_sheet.write(row,19,app_no_submit)
                excel_file_write.save("Database.xls")
                clearing()
                table_contents()
            
        else:
            msg.showinfo("Information","Application Number Does Not Exists")
    except IndexError:
        msg.showwarning("Warning","Please Select A Student Detail From The Table")
    return

def submit():
    fst_name_submit=(fst_name.get())
    initial_submit=initial.get()
    dob_date_submit=dob_date.get()
    dob_month_submit=dob_month.get()
    dob_year_submit=dob_year.get()
    age_submit=age.get()
    gender_submit=gender.get()
    father_name_submit=father_name.get()
    father_mobile_submit=father_mobile.get()
    mother_name_submit=mother_name.get()
    mother_mobile_submit=mother_mobile.get()
    whatsapp_number_submit=whatsapp_number.get()
    house_no_submit=house_no.get()
    address_submit=address.get()
    district_submit=district.get()
    pin_code_submit=pin_code.get()
    state_submit=state.get()
    father_mail_submit=father_mail.get()
    student_mail_submit=student_mail.get()
    group_submit=group.get()
    section_submit=section.get()
    app_no_submit=app_no.get()


    def submitting():
        row=0
        font_hedding=Font()
        font_hedding.name="Arial Black"

        alignment_hedding=Alignment()
        alignment_hedding.horizontal="CENTER"
        alignment_hedding.vertical="CENTER"

        border_hedding=Borders()
        border_hedding.left=Borders.THICK
        border_hedding.right=Borders.THICK
        border_hedding.top=Borders.THICK
        border_hedding.bottom=Borders.THICK

        pattern_hedding=Pattern()
        pattern_hedding.pattern=Pattern.SOLID_PATTERN
        pattern_hedding.pattern_fore_colour=0x9A1F

        style_hedding=XFStyle()
        style_hedding.font=font_hedding
        style_hedding.alignment=alignment_hedding
        style_hedding.borders=border_hedding
        style_hedding.pattern=pattern_hedding
        def creating():
            excel_file=Workbook()
            excel_sheet=excel_file.add_sheet("Data")
            excel_sheet.write(0,0,"S\\no",style_hedding)
            excel_sheet.col(0).width=1400
            excel_sheet.write(0,1,"Name",style_hedding)
            excel_sheet.col(1).width=5000
            excel_sheet.write(0,2,"Initial",style_hedding)
            excel_sheet.col(2).width=1800
            excel_sheet.write(0,3,"Date Of Birth",style_hedding)
            excel_sheet.col(3).width=4000
            excel_sheet.write(0,4,"Age",style_hedding)
            excel_sheet.col(4).width=1300
            excel_sheet.write(0,5,"Gender",style_hedding)
            excel_sheet.col(5).width=2000
            excel_sheet.write(0,6,"Father Name",style_hedding)
            excel_sheet.col(6).width=4000
            excel_sheet.write(0,7,"Father Mobile",style_hedding)
            excel_sheet.col(7).width=4500
            excel_sheet.write(0,8,"Mother Name",style_hedding)
            excel_sheet.col(8).width=4000
            excel_sheet.write(0,9,"Mother Mobile",style_hedding)
            excel_sheet.col(9).width=4500
            excel_sheet.write(0,10,"Whatsapp Number",style_hedding)
            excel_sheet.col(10).width=5100
            excel_sheet.write(0,11,"Address",style_hedding)
            excel_sheet.col(11).width=8000
            excel_sheet.write(0,12,"District",style_hedding)
            excel_sheet.col(12).width=3000
            excel_sheet.write(0,13,"State",style_hedding)
            excel_sheet.col(13).width=3000
            excel_sheet.write(0,14,"Pin Code",style_hedding)
            excel_sheet.col(14).width=2600
            excel_sheet.write(0,15,"Father Mail",style_hedding)
            excel_sheet.col(15).width=5000
            excel_sheet.write(0,16,"Student Mail",style_hedding)
            excel_sheet.col(16).width=5000
            excel_sheet.write(0,17,"Group",style_hedding)
            excel_sheet.col(17).width=2000
            excel_sheet.write(0,18,"Section",style_hedding)
            excel_sheet.col(18).width=2300
            excel_sheet.write(0,19,"Application Number",style_hedding)
            excel_sheet.col(19).width=5500
            excel_file.save("Database.xls")
        try:
            excel_file=xlrd.open_workbook("Database.xls")
            for sheet in excel_file.sheets():
                for rows_in_sheet in range(sheet.nrows):
                    row+=1
            application_no=0
            column=19
            app_no_submit_checking0=app_no_submit
            for sheet in excel_file.sheets():
                for row_in_sheet in range(sheet.nrows):
                    for column_in_sheet in range(sheet.ncols):
                        if column_in_sheet==column:
                            app_no_excel=sheet.cell(row_in_sheet,column_in_sheet).value
                            if app_no_submit_checking0==app_no_excel:
                                application_no+=1
                                msg.showwarning("Warning","Application Number Already Exists")
                                break               
            if application_no==0:
                excel_file_read_write=xlrd.open_workbook("Database.xls")
                excel_file_write=copy(excel_file_read_write)
                excel_file_write_sheet=excel_file_write.get_sheet(0)
                excel_file_write_sheet.write(row,0,row,style_hedding)
                excel_file_write_sheet.write(row,1,fst_name_submit,style_hedding)
                excel_file_write_sheet.write(row,2,initial_submit,style_hedding)
                excel_file_write_sheet.write(row,3,dob_date_submit+" "+dob_month_submit+" "+dob_year_submit,style_hedding)
                excel_file_write_sheet.write(row,4,age_submit,style_hedding)
                excel_file_write_sheet.write(row,5,gender_submit,style_hedding)
                excel_file_write_sheet.write(row,6,father_name_submit,style_hedding)
                excel_file_write_sheet.write(row,7,father_mobile_submit,style_hedding)
                excel_file_write_sheet.write(row,8,mother_name_submit,style_hedding)
                excel_file_write_sheet.write(row,9,mother_mobile_submit,style_hedding)
                excel_file_write_sheet.write(row,10,whatsapp_number_submit,style_hedding)
                excel_file_write_sheet.write(row,11,house_no_submit+" "+address_submit,style_hedding)
                excel_file_write_sheet.write(row,12,district_submit,style_hedding)
                excel_file_write_sheet.write(row,13,pin_code_submit,style_hedding)
                excel_file_write_sheet.write(row,14,state_submit,style_hedding)
                excel_file_write_sheet.write(row,15,father_mail_submit,style_hedding)
                excel_file_write_sheet.write(row,16,student_mail_submit,style_hedding)
                excel_file_write_sheet.write(row,17,group_submit,style_hedding)
                excel_file_write_sheet.write(row,18,section_submit,style_hedding)
                excel_file_write_sheet.write(row,column,app_no_submit,style_hedding)
                excel_file_write_sheet.col(column).width=5500
                excel_file_write.save("Database.xls")
                msg.showinfo("Information","The Details Of The Student Has Been Successfully Noted")
                ask=msg.askyesno("Information","Click Yes To Upload The Image Of The Student")
                if ask==True:
                    image_ftn()
                else:
                    ask2=msg.askyesnocancel("Warning","Click Yes To Add Image Otherwise You Wont't Be Able To Add It")
                    if ask2==True:
                        image_ftn()
                    else:
                        pass
                clearing()       
        except FileNotFoundError:
            creating_database=msg.askyesno("Welcome","You Are New To This Software \nClick Yes To Create A New Database")
            if creating_database==True:
                creating()
                submitting()
            else:
                pass


    for checking in range(1):
        if fst_name_submit=="":
            msg.showwarning("Warning","First Name Column Is Mandatory.")
            break
        if initial_submit=="":
            msg.showwarning("Warning","Last Name Column Is Mandatory.")
            break
        if dob_date_submit=="":
            msg.showwarning("Warning","Date Of Birth \"Date\" Column Is Mandatory.")
            break
        if dob_month_submit=="":
            msg.showwarning("Warning","Date Of Birth \"Month\" Column Is Mandatory")
            break
        if dob_year_submit=="":
            msg.showwarning("Warning","Date Of Birth \"Year\" Column Is Mandatory")
            break
        if age_submit=="":
            msg.showwarning("Warning","Age Column Is Mandatory.")
            break
        if gender_submit=="":
            msg.showwarning("Warning","Gender Column Is Mandatory.")
            break
        if father_name_submit=="":
            msg.showwarning("Warning","Father Name Column Is Mandatory.")
            break
        if father_mobile_submit=="":
            msg.showwarning("Warning","Father Mobile Number Column Is Mandatory")
            break
        if mother_name_submit=="":
            msg.showwarning("Warning","Mother Name Column Is Mandatory")
            break
        if mother_mobile_submit=="":
            msg.showwarning("Warning","Mother Mobile Number Column Is Mandatory")
            break
        if whatsapp_number_submit=="":
            msg.showwarning("Warning","Whatsapp Number Column Is Mandatory.")
            break
        if house_no_submit=="":
            msg.showwarning("Warning","House Number Column Is Mandatory.")
            break
        if address_submit=="":
            msg.showwarning("Warning","Address Number Column Is Mandatory")
            break
        if district_submit=="":
            msg.showwarning("Warning","District Column Is Mandatory")
            break
        if pin_code_submit=="":
            msg.showwarning("Warning","Pin Code Column Is Mandatory")
            break

        if state_submit=="":
            msg.showwarning("Warning","State Column Is Mandatory")
            break
        if father_mail_submit=="":
            msg.showwarning("Warning","Father Mail I'd Column Is Mandatory.")
            break
        if student_mail_submit=="":
            msg.showwarning("Warning","Student Mail I'D Column Is Mandatory.")
            break
        if group_submit=="":
            msg.showwarning("Warning","Group Number Column Is Mandatory")
            break
        if section_submit=="":
            msg.showwarning("Warning","Section Column Is Mandatory")
            break
        if app_no_submit=="":
            msg.showwarning("Warning","Application Number Column Is Mandatory")
            break
        if "CHECKED"=="CHECKED":
            submitting()
            table_contents()
                    

    return
    

def search():
    search_key_user=search_key.get()
    search_value=value.get()
    
    def searching(column_no,column_value,column_name):
        try:
            open_excel=xlrd.open_workbook("Database.xls")
            table.delete(*table.get_children())
            for sheets_excel in open_excel.sheets():
                for row_sheet in range(sheets_excel.nrows):
                    for col_sheet in range(sheets_excel.ncols):
                        if col_sheet==column_no:
                            
                            row_sheet_value=row_sheet
                            search_value_excel=sheets_excel.cell(row_sheet,col_sheet).value
                            
                            if search_value_excel==column_value:
                                search_row_list=sheets_excel.row_values(row_sheet)
                                search_0=int(search_row_list[0])
                                del search_row_list[0]
                                search_row_list.insert(0,search_0)
                                a=tuple(search_row_list)
                                b=list(a)                                               
                                table.insert("","end",values=b)
                                
        except FileNotFoundError:
            pass
    if search_key_user=="Name":
        column_name=search_key_user
        column_no=1
        column_value=search_value
        searching(column_no,column_value,column_name)
    elif search_key_user=="Application Number":
        column_name=search_key_user
        column_no=19
        column_value=search_value
        searching(column_no,column_value,column_name)
    elif search_key_user=="Group":
        column_name=search_key_user
        column_no=17
        column_value=search_value
        searching(column_no,column_value,column_name)
    elif search_key_user=="Section":
        column_name=search_key_user
        column_no=18
        column_value=search_value
        searching(column_no,column_value,column_name)
    elif search_key_user=="Group & Section Example: 2 A":
        column_no1=17
        column_no2=18
        GROUP=search_value[0]
        SECTION=search_value[2]
        open_excel1=xlrd.open_workbook("Database.xls")
        table.delete(*table.get_children())
        for sheets_excel1 in open_excel1.sheets():
            for row_sheet1 in range(sheets_excel1.nrows):
                for col_sheet1 in range(sheets_excel1.ncols):
                    if col_sheet1==column_no1:
                        row_sheet_value1=row_sheet1
                        search_value_excel1=sheets_excel1.cell(row_sheet1,col_sheet1).value
                        if search_value_excel1==GROUP:
                            col_sheet2=col_sheet1+1
                            search_value_excel2=sheets_excel1.cell(row_sheet1,col_sheet2).value
                            if search_value_excel2==SECTION:
                                search_row_list=sheets_excel1.row_values(row_sheet1)
                                search_0=int(search_row_list[0])
                                del search_row_list[0]
                                search_row_list.insert(0,search_0)
                                a=tuple(search_row_list)
                                b=list(a)                                               
                                table.insert("","end",values=b)
                                

                            
    return

#Opening A GUI Page
database=tk.Tk()
database.title("Student DataBase Management")
database.geometry("1350x700+0+0")
#database.iconbitmap(r"Icons\Student Database.ico")
database.configure(background="aqua")
#Hedding
hedding=ttk.Label(text="Student Database Management",border=5,background="yellow",foreground="red",font=("jokerman",20))
hedding.place(x=490,y=1)
#hedding.pack(side="top",fill="x")

#Parent Variables
table_frame=ttk.Frame(database,border=4)
table_frame.place(x=768,y=46,width=591,height=663)

table_details_frame=ttk.Frame(table_frame,border=10)
table_details_frame.place(x=6,y=58,width=585,height=603)

new_std=ttk.LabelFrame(database,text="Mandatory Details",)
new_std.place(y=46,x=10,height=663)

parent_det=ttk.LabelFrame(new_std,text="Parents' Details")
parent_det.grid(row=6,column=0,padx=8,pady=6,columnspan=8)

communication=ttk.LabelFrame(new_std,text="Communication Details")
communication.grid(row=7,column=0,padx=8,pady=6,columnspan=8)

office_det=ttk.LabelFrame(new_std,text="Office Details",)
office_det.grid(row=8,column=0,padx=8,pady=6,columnspan=4)

#Creating The Main Page

fst_name_lbl=ttk.Label(new_std,text="First Name :",foreground="brown",font=("forte",15)).grid(row=2,column=0,sticky="W")
fst_name=tk.StringVar()
fst_name_entry=ttk.Entry(new_std,textvariable=fst_name,width=20,font=("hobostd",11))
fst_name_entry.grid(row=2,column=1,columnspan=2,sticky="W")

initial_lbl=ttk.Label(new_std,text="Initial :",foreground="brown",font=("forte",15)).grid(row=2,column=3,padx=5)
initial=tk.StringVar()
initial_entry=ttk.Entry(new_std,textvariable=initial,width=20,font=("hobostd",11))
initial_entry.grid(row=2,column=4)

dob_lbl=tk.Label(new_std,text="Date Of Birth :",foreground="brown",font=("forte",15)).grid(column=0,row=3,sticky="W",pady=5)
dob_date=tk.StringVar()
dob_month=tk.StringVar()
dob_year=tk.StringVar()
date=[]
month=["January","February","March","April","May","June","July","August","September","October","November","December"]
year=[]
for dates in range(1,32,1):
    date.append(dates)
for years in range(1990,2016,1):
    year.append(years)
dob_date_entry=ttk.Combobox(new_std,width=3,textvariable=dob_date,font=("hobostd",11))
dob_date_entry["values"]=(date)
dob_date_entry.grid(column=1,row=3,sticky="W",pady=5)
dob_month_entry=ttk.Combobox(new_std,width=7,textvariable=dob_month,font=("hobostd",11))
dob_month_entry["values"]=(month)
dob_month_entry.grid(column=2,row=3,pady=5)
dob_year_entry=ttk.Combobox(new_std,width=5,textvariable=dob_year,font=("hobostd",11))
dob_year_entry["values"]=(year)
dob_year_entry.grid(column=3,row=3,pady=5)

age_lbl=ttk.Label(new_std,text="Age :",foreground="brown",font=("forte",15)).grid(row=4,column=0,sticky="W",pady=5)
age=tk.StringVar()
age_combobox=ttk.Combobox(new_std,textvariable=age,width=5,font=("hobostd",11))
age_combobox["values"]=["5","6","7","8","9","10","11","12","13","15","16","17","18","19","20","21","22","23","24","25"]
age_combobox.grid(row=4,column=1,sticky="W",pady=5)

gender_lbl=ttk.Label(new_std,text="Genger :",foreground="brown",font=("forte",15)).grid(row=5,column=0,sticky="W",pady=5)
gender=tk.StringVar()
gender_combobox=ttk.Combobox(new_std,textvariable=gender,width=10,font=("hobostd",11))
gender_combobox["values"]=["Male","Female","Transgender"]
gender_combobox.grid(row=5,column=1,sticky="W",pady=5)

father_name_lbl=ttk.Label(parent_det,text="Fathers's Name :",foreground="purple",font=("forte",15)).grid(row=0,column=0,pady=5)
father_name=tk.StringVar()
father_name_entry=ttk.Entry(parent_det,textvariable=father_name,font=("hobostd",11))
father_name_entry.grid(row=0,column=1,pady=5)

father_mobile_lbl=ttk.Label(parent_det,text="Mobile Number :",foreground="green",font=("forte",15)).grid(row=0,column=3,pady=5,padx=10)
father_mobile=tk.StringVar()
father_mobile_entry=ttk.Entry(parent_det,textvariable=father_mobile,font=("hobostd",11))
father_mobile_entry.grid(row=0,column=4,pady=5)
father_mobile.set("+91 ")

mother_name_lbl=ttk.Label(parent_det,text="Mother's Name :",foreground="purple",font=("forte",15)).grid(row=1,column=0,pady=5)
mother_name=tk.StringVar()
mother_name_entry=ttk.Entry(parent_det,textvariable=mother_name,font=("hobostd",11))
mother_name_entry.grid(column=1,row=1,pady=5)

mother_mobile_lbl=ttk.Label(parent_det,text="Mobile Number :",foreground="green",font=("forte",15)).grid(row=1,column=3,pady=5,padx=10)
mother_mobile=tk.StringVar()
mother_mobile_entry=ttk.Entry(parent_det,textvariable=mother_mobile,font=("hobostd",11))
mother_mobile_entry.grid(row=1,column=4,pady=5)
mother_mobile.set("+91 ")

whatsapp_number_lbl=ttk.Label(parent_det,text="Whatsapp Number :",foreground="green",font=("forte",15)).grid(row=2,column=1,pady=5)
whatsapp_number=tk.StringVar()
whatsapp_number_entry=ttk.Entry(parent_det,textvariable=whatsapp_number,font=("hobostd",11))
whatsapp_number_entry.grid(row=2,column=2,pady=5,columnspan=2)
whatsapp_number.set("+91 ")

house_no_lbl=ttk.Label(communication,text="House \ Appartement No. :",foreground="#FF1166",font=("forte",15)).grid(row=0,column=0,pady=5)
house_no=tk.StringVar()
house_no_entry=ttk.Entry(communication,textvariable=house_no,width=10,font=("hobostd",11))
house_no_entry.grid(row=0,column=1,sticky="W",pady=5)

address_lbl=ttk.Label(communication,text="Address :",foreground="#FF1166",font=("forte",15)).grid(row=1,column=0,sticky="W",pady=5)
address=tk.StringVar()
address_entry=ttk.Entry(communication,textvariable=address,width=40,font=("hobostd",11))
address_entry.grid(row=1,column=1,sticky="W",columnspan=5,pady=5)

district_lbl=ttk.Label(communication,text="District :",foreground="#FF1166",font=("forte",15)).grid(row=2,column=0,sticky="W",pady=5)
district=tk.StringVar()
district_entry=ttk.Entry(communication,textvariable=district,width=15,font=("hobostd",11))
district_entry.grid(row=2,column=1,sticky="W",pady=5)

pin_code_lbl=ttk.Label(communication,text="Pin Code :",foreground="#FF1166",font=("forte",15)).grid(row=2,column=3,pady=5)
pin_code=tk.StringVar()
pin_code_entry=ttk.Entry(communication,textvariable=pin_code,width=8,font=("hobostd",11))
pin_code_entry.grid(row=2,column=4,pady=5,sticky="W")

state_lbl=ttk.Label(communication,text="State :",foreground="#FF1166",font=("forte",15)).grid(row=3,column=0,sticky="W",pady=5)
state=tk.StringVar()
state_entry=ttk.Entry(communication,textvariable=state,font=("hobostd",11))
state_entry.grid(row=3,column=1,sticky="W")

father_mail=ttk.Label(communication,text="Mail :",foreground="#FF1166",font=("forte",15)).grid(sticky="E",padx=10,row=4,column=0,pady=5)
father_mail=tk.StringVar()
father_mail_entry=ttk.Entry(communication,width=30,textvariable=father_mail,font=("hobostd",11))
father_mail_entry.grid(row=4,column=1,pady=5,columnspan=2)
father_mail_info=ttk.Label(communication,text="(Father)",foreground="#FF1166",font=("forte",15)).grid(row=4,column=3,pady=5,padx=10)

student_mail=ttk.Label(communication,text="Mail :",foreground="#FF1166",font=("forte",15)).grid(sticky="E",padx=10,row=5,column=0,pady=5)
student_mail=tk.StringVar()
student_mail_entry=ttk.Entry(communication,width=30,textvariable=student_mail,font=("hobostd",11))
student_mail_entry.grid(row=5,column=1,pady=5,columnspan=2)
student_mail_info=ttk.Label(communication,text="(Student)",foreground="#FF1166",font=("forte",15)).grid(row=5,column=3,pady=5,padx=10)

group_lbl=ttk.Label(office_det,text="Group :",foreground="indigo",font=("forte",15))
group_lbl.grid(row=0,column=0,sticky="W",pady=8)
group=tk.StringVar()
group_entry=ttk.Combobox(office_det,textvariable=group,font=("hobostd",11),width=3)
group_entry["values"]=["I","II","III"]
group_entry.grid(row=0,column=1,padx=10)

section_lbl=ttk.Label(office_det,text="Section :",foreground="indigo",font=("forte",15))
section_lbl.grid(row=0,column=3,sticky="W",padx=10)
section=tk.StringVar()
section_entry=ttk.Combobox(office_det,textvariable=section,font=("hobostd",11),width=2)
section_entry["values"]=["A","B","C"]
section_entry.grid(row=0,column=4,padx=10)

app_no_lbl=ttk.Label(office_det,text="Application No. :",foreground="indigo",font=("forte",15)).grid(pady=8,row=0,sticky="W",padx=10,column=5)
app_no=tk.IntVar()
app_no_entry=ttk.Entry(office_det,width=5,textvariable=app_no,font=("hobostd",11))
app_no_entry.grid(row=0,column=6,padx=10)

submit_btn=ttk.Button(new_std,text="Submit",command=submit).grid(row=8,column=4)

update_btn=ttk.Button(new_std,text="Update",command=update).grid(row=9,column=1,pady=8,columnspan=3)

open_excel_btn=ttk.Button(new_std,text="Clear All",command=clear).grid(row=9,column=0,pady=8,columnspan=2)

exit_software_btn=ttk.Button(new_std,text="Exit",command=exit_software).grid(row=9,column=3,pady=8,columnspan=2)

search_lbl=ttk.Label(table_frame,text="Search By :",foreground="orange",font=("hobostd",15)).grid(row=0,column=0)

search_key=tk.StringVar()
search_entry=ttk.Combobox(table_frame,textvariable=search_key,foreground="black",font=("forte",14),width=15)
search_entry["values"]=["Name","Application Number","Group","Section","Group & Section Example: 2 A"]
search_entry.grid(row=0,column=1,sticky="W",padx=5)

value_lbl=ttk.Label(table_frame,text="Value:",foreground="orange",font=("hobostd",15)).grid(row=0,column=2)

value=tk.StringVar()
value_entry=ttk.Entry(table_frame,textvariable=value,foreground="black",font=("forte",14),width=15).grid(row=0,column=3)

search_btn=ttk.Button(table_frame,text="Search",command=search).grid(row=0,column=4,padx=5,sticky="W")

showall_btn=ttk.Button(table_frame,text="Show All",command=table_contents).grid(row=1,column=0,padx=10)

scroll_x=ttk.Scrollbar(table_details_frame,orient="horizontal")
scroll_y=ttk.Scrollbar(table_details_frame,orient="vertical")
table=ttk.Treeview(table_details_frame,columns=("serial_no","name","initial","date_of_birth","age","gender","father_name","father_mobile","mother_name","mother_mobile","whatsapp_no","address","district","state","pin_code","father_mail","student_mail","group","section","application_no"),xscrollcommand=scroll_x.set,yscrollcommand=scroll_y.set)

scroll_x.pack(side="bottom",fill="x")
scroll_y.pack(side="right",fill="y")
scroll_x.config(command=table.xview)
scroll_y.config(command=table.yview)
table.heading("serial_no",text="S\\no")
table.heading("name",text="Name")
table.heading("initial",text="Initial")
table.heading("date_of_birth",text="Date-Of-Birth")
table.heading("age",text="Age")
table.heading("gender",text="Gender")
table.heading("father_name",text="Father Name")
table.heading("father_mobile",text="Father Mobile")
table.heading("mother_name",text="Mother Name")
table.heading("mother_mobile",text="Mother Mobile")
table.heading("whatsapp_no",text="Whatsapp Number")
table.heading("address",text="Address")
table.heading("district",text="District")
table.heading("state",text="State")
table.heading("pin_code",text="Pin Code")
table.heading("father_mail",text="Father Mail")
table.heading("student_mail",text="Student Mail")
table.heading("group",text="Group")
table.heading("section",text="Section")
table.heading("application_no",text="Application Number")

table["show"]="headings"
table.column("serial_no",width=40)
table.column("name",width=150)
table.column("initial",width=50)
table.column("date_of_birth",width=100)
table.column("age",width=40)
table.column("gender",width=50)
table.column("father_name",width=150)
table.column("father_mobile",width=100)
table.column("mother_name",width=150)
table.column("mother_mobile",width=100)
table.column("whatsapp_no",width=110)
table.column("address",width=300)
table.column("district",width=45)
table.column("state",width=70)
table.column("pin_code",width=70)
table.column("father_mail",width=150)
table.column("student_mail",width=150)
table.column("group",width=42)
table.column("section",width=50)
table.column("application_no",width=120)
table.pack(fill="both",expand=1)

table.bind("<ButtonRelease-1>",get_cursor)
table_contents()
#Forcing to Start The GUI
database.mainloop()
